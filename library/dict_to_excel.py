#!/usr/bin/python

from ansible.module_utils.basic import *
from openpyxl import Workbook,load_workbook
from distutils.util import strtobool

def dict_to_excel(col,data,save_fn,newline=True):
    # get data to enter to exisiting excel file and insert it in the next free raw
    # col - column to insert the data too 
    # data - data to insert to excel 
    # save_fn - file to be saved 
    # newline - to insert the data in a new line or in the last line 
    wb=load_workbook(filename=save_fn)
    sheet=wb.active
    if newline == True:
        sheet. cell(sheet.max_row+1,col,data)
    else:
          sheet.cell(sheet.max_row,col,data)  
    wb.save(filename=save_fn)
   # return "data was saved"

def main():

    fields = {
        "column": {"default": True, "type": "int"},
        "data": {"default": True, "type": "str"},
        "file_to_save": {"default": True, "type": "str"},
        "newline":{"default": True,"type":"bool"}
    }


    module = AnsibleModule(argument_spec=fields)
    dict_to_excel(module.params["column"],module.params["data"],module.params["file_to_save"],module.params["newline"])
    theReturnValue = {"status": "data was saved"}
    module.exit_json(changed=False, meta=theReturnValue)

if __name__ == '__main__':
    main()