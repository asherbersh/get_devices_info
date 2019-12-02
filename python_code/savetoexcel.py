from openpyxl import Workbook,load_workbook

def data_to_excel(col,data,save_fn,newline=True):
    # get data to enter to exisiting excel file and insert it in the next free raw
    # col - column to insert the data too 
    # data - data to insert to excel 
    # save_fn - file to be saved 
    # newline - to insert the data in a new line or in the last line 
    wb=load_workbook(filename=save_fn)
    sheet=wb.active
    if newline == True:
        sheet. cell(sheet.max_row+1,col,data)
    else:
          sheet.cell(sheet.max_row,col,data)  
    wb.save(filename=save_fn)
    return "data was saved"

def dict_to_excel(dict_data,save_fn,newline=True):
    # get data to enter to exisiting excel file and insert it in the next free raw
    # dict_data - data to insert to excel that include column and data to be insert 
    # save_fn - file to be saved 
    # newline - to insert the data in a new line or in the last line 
    wb=load_workbook(filename=save_fn)
    sheet=wb.active


    if newline == True:
        insert_raw=sheet.max_row+1
        for col,data in dict_data.items():
            sheet. cell(insert_raw,col,data)
    else:
        insert_raw=sheet.max_row
        for col,data in dict_data.items():
          sheet.cell(insert_raw,col,data)  
    wb.save(filename=save_fn)
    return "data was saved"




#print (data_to_excel(1,"floor8_ptsw1","device_report.xlsx"))
#print (data_to_excel(2,"192.168.0.3","device_report.xlsx",False))
#print (data_to_excel(3,"15.0","device_report.xlsx",False))

dict={1:"asher1",2:"the men1"}
print (dict_to_excel(dict,"device_report.xlsx",False))
